from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import connection
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
import json
import re
import csv
import io
from django.template.loader import render_to_string
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError) as e:
    print(f"⚠️ WeasyPrint no disponible: {e}")
    WEASYPRINT_AVAILABLE = False
    HTML = None
    CSS = None

try:
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .models import ReportTemplate, GeneratedReport, VoiceQuery
from .serializers import ReportTemplateSerializer, GeneratedReportSerializer, VoiceQuerySerializer
from sales.models import Order, OrderItem
from products.models import Product
from users.models import User

class ReportTemplateViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para plantillas de reportes predefinidos"""
    queryset = ReportTemplate.objects.filter(is_active=True)
    serializer_class = ReportTemplateSerializer
    permission_classes = [IsAuthenticated]

class ReportsViewSet(viewsets.ModelViewSet):
    """ViewSet principal para generación de reportes"""
    serializer_class = GeneratedReportSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return GeneratedReport.objects.filter(user=self.request.user)
    
    @action(detail=False, methods=['post'])
    def generate_predefined(self, request):
        """Generar reporte predefinido"""
        template_id = request.data.get('template_id')
        parameters = request.data.get('parameters', {})
        
        try:
            template = ReportTemplate.objects.get(id=template_id)
            
            # Generar datos según el tipo de reporte
            if template.category == 'SALES':
                data = self._generate_sales_report(parameters)
            elif template.category == 'INVENTORY':
                data = self._generate_inventory_report(parameters)
            elif template.category == 'CUSTOMERS':
                data = self._generate_customers_report(parameters)
            else:
                data = {'error': 'Tipo de reporte no soportado'}
            
            # Crear registro del reporte
            report = GeneratedReport.objects.create(
                user=request.user,
                template=template,
                title=template.name,
                description=template.description,
                query_text=f"Reporte predefinido: {template.name}",
                query_sql="Generated from template",
                parameters=parameters,
                data=data,
                status='COMPLETED',
                completed_at=timezone.now()
            )
            
            return Response({
                'id': report.id,
                'title': report.title,
                'data': data,
                'created_at': report.created_at,
                'status': report.status
            })
            
        except ReportTemplate.DoesNotExist:
            return Response(
                {'error': 'Plantilla de reporte no encontrada'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def generate_custom(self, request):
        """Generar reporte personalizado con IA"""
        query_text = request.data.get('query_text', '')
        
        if not query_text:
            return Response(
                {'error': 'Consulta requerida'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Interpretar consulta con IA simple
            interpreted_query = self._interpret_query(query_text)
            data = self._execute_interpreted_query(interpreted_query)
            
            # Crear registro del reporte
            report = GeneratedReport.objects.create(
                user=request.user,
                title=f"Reporte personalizado: {query_text[:50]}...",
                query_text=query_text,
                query_sql=interpreted_query.get('sql', ''),
                parameters=interpreted_query.get('parameters', {}),
                data=data,
                status='COMPLETED',
                completed_at=timezone.now()
            )
            
            return Response({
                'id': report.id,
                'title': report.title,
                'data': data,
                'interpretation': interpreted_query,
                'created_at': report.created_at
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar reporte: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def process_voice_query(self, request):
        """Procesar consulta de voz"""
        transcribed_text = request.data.get('transcribed_text', '')
        
        if not transcribed_text:
            return Response(
                {'error': 'Texto transcrito requerido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Interpretar consulta de voz
            interpreted_query = self._interpret_query(transcribed_text)
            data = self._execute_interpreted_query(interpreted_query)
            
            # Crear reporte
            report = GeneratedReport.objects.create(
                user=request.user,
                title=f"Consulta por voz: {transcribed_text[:50]}...",
                query_text=transcribed_text,
                query_sql=interpreted_query.get('sql', ''),
                data=data,
                status='COMPLETED',
                completed_at=timezone.now()
            )
            
            # Crear registro de consulta de voz
            voice_query = VoiceQuery.objects.create(
                user=request.user,
                transcribed_text=transcribed_text,
                interpreted_query=json.dumps(interpreted_query),
                confidence_score=0.95,  # Simulado
                generated_report=report
            )
            
            return Response({
                'id': report.id,
                'title': report.title,
                'data': data,
                'voice_query_id': voice_query.id,
                'interpretation': interpreted_query
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar consulta de voz: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _generate_sales_report(self, parameters):
        """Generar reporte de ventas"""
        # Parámetros por defecto
        days = parameters.get('days', 30)
        start_date = timezone.now() - timedelta(days=days)
        
        # Consultar ventas
        orders = Order.objects.filter(
            created_at__gte=start_date,
            status__in=['DELIVERED', 'PAID']
        )
        
        # Calcular métricas
        total_sales = sum(float(order.total) for order in orders)
        total_orders = orders.count()
        avg_order_value = total_sales / total_orders if total_orders > 0 else 0
        
        # Productos más vendidos
        top_products = {}
        for order in orders:
            for item in order.items.all():
                product_name = item.product.name
                if product_name in top_products:
                    top_products[product_name] += item.quantity
                else:
                    top_products[product_name] = item.quantity
        
        top_products_list = sorted(top_products.items(), key=lambda x: x[1], reverse=True)[:10]
        
        # Ventas por día
        daily_sales = {}
        for order in orders:
            date_key = order.created_at.strftime('%Y-%m-%d')
            if date_key in daily_sales:
                daily_sales[date_key] += float(order.total)
            else:
                daily_sales[date_key] = float(order.total)
        
        return {
            'period': f'Últimos {days} días',
            'total_sales': total_sales,
            'total_orders': total_orders,
            'average_order_value': round(avg_order_value, 2),
            'top_products': [{'name': name, 'quantity': qty} for name, qty in top_products_list],
            'daily_sales': [{'date': date, 'amount': amount} for date, amount in sorted(daily_sales.items())],
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_inventory_report(self, parameters):
        """Generar reporte de inventario"""
        products = Product.objects.all()
        
        low_stock_threshold = parameters.get('low_stock_threshold', 10)
        
        inventory_data = []
        low_stock_items = []
        total_value = 0
        
        for product in products:
            stock_value = float(product.price) * product.stock
            total_value += stock_value
            
            product_data = {
                'name': product.name,
                'sku': product.sku,
                'stock': product.stock,
                'price': float(product.price),
                'stock_value': stock_value,
                'category': product.category.name if product.category else 'Sin categoría'
            }
            
            inventory_data.append(product_data)
            
            if product.stock <= low_stock_threshold:
                low_stock_items.append(product_data)
        
        return {
            'total_products': len(inventory_data),
            'total_inventory_value': round(total_value, 2),
            'low_stock_threshold': low_stock_threshold,
            'low_stock_items': low_stock_items,
            'inventory_details': inventory_data,
            'generated_at': timezone.now().isoformat()
        }
    
    def _generate_customers_report(self, parameters):
        """Generar reporte de clientes"""
        users = User.objects.filter(is_staff=False)
        
        customers_data = []
        for user in users:
            orders = Order.objects.filter(user=user)
            total_spent = sum(float(order.total) for order in orders)
            
            customers_data.append({
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'total_orders': orders.count(),
                'total_spent': total_spent,
                'last_order': orders.last().created_at.isoformat() if orders.exists() else None,
                'date_joined': user.date_joined.isoformat()
            })
        
        # Ordenar por total gastado
        customers_data.sort(key=lambda x: x['total_spent'], reverse=True)
        
        return {
            'total_customers': len(customers_data),
            'customers': customers_data,
            'generated_at': timezone.now().isoformat()
        }
    
    def _interpret_query(self, query_text):
        """Interpretar consulta en lenguaje natural (IA simple)"""
        query_lower = query_text.lower()
        
        # Patrones simples de reconocimiento
        if any(word in query_lower for word in ['ventas', 'vendido', 'ingresos', 'facturación']):
            if 'mes' in query_lower or 'mensual' in query_lower:
                days = 30
            elif 'semana' in query_lower or 'semanal' in query_lower:
                days = 7
            elif 'año' in query_lower or 'anual' in query_lower:
                days = 365
            else:
                days = 30
            
            return {
                'type': 'sales',
                'parameters': {'days': days},
                'sql': f'Sales report for last {days} days'
            }
        
        elif any(word in query_lower for word in ['inventario', 'stock', 'productos']):
            return {
                'type': 'inventory',
                'parameters': {},
                'sql': 'Inventory report'
            }
        
        elif any(word in query_lower for word in ['clientes', 'usuarios', 'compradores']):
            return {
                'type': 'customers',
                'parameters': {},
                'sql': 'Customers report'
            }
        
        else:
            return {
                'type': 'general',
                'parameters': {},
                'sql': 'General query'
            }
    
    def _execute_interpreted_query(self, interpreted_query):
        """Ejecutar consulta interpretada"""
        query_type = interpreted_query.get('type')
        parameters = interpreted_query.get('parameters', {})
        
        if query_type == 'sales':
            return self._generate_sales_report(parameters)
        elif query_type == 'inventory':
            return self._generate_inventory_report(parameters)
        elif query_type == 'customers':
            return self._generate_customers_report(parameters)
        else:
            return {'message': 'Consulta no reconocida', 'query': interpreted_query}

    @action(detail=False, methods=['post'])
    def export_pdf(self, request):
        """Exportar reporte a PDF"""
        try:
            report_id = request.data.get('report_id')
            if not report_id:
                return Response({'error': 'report_id es requerido'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                report = GeneratedReport.objects.get(id=report_id)
            except GeneratedReport.DoesNotExist:
                return Response({'error': 'Reporte no encontrado'}, status=status.HTTP_404_NOT_FOUND)
            
            # Generar HTML para el PDF
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <title>Reporte - {report.title}</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        margin: 20px;
                        color: #333;
                    }}
                    .header {{
                        background: linear-gradient(135deg, #00BCD4, #0097A7);
                        color: white;
                        padding: 20px;
                        margin-bottom: 20px;
                        border-radius: 8px;
                    }}
                    .header h1 {{
                        margin: 0;
                        font-size: 24px;
                    }}
                    .header p {{
                        margin: 5px 0 0 0;
                        opacity: 0.9;
                    }}
                    .content {{
                        background: #f8f9fa;
                        padding: 20px;
                        border-radius: 8px;
                        border: 1px solid #e9ecef;
                    }}
                    .data-table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin-top: 20px;
                    }}
                    .data-table th, .data-table td {{
                        border: 1px solid #ddd;
                        padding: 12px;
                        text-align: left;
                    }}
                    .data-table th {{
                        background-color: #00BCD4;
                        color: white;
                    }}
                    .footer {{
                        margin-top: 30px;
                        text-align: center;
                        color: #666;
                        font-size: 12px;
                    }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>📊 {report.title}</h1>
                    <p>Generado el: {report.created_at.strftime('%d/%m/%Y %H:%M')}</p>
                    <p>Estado: {report.status}</p>
                </div>
                
                <div class="content">
                    <h2>Datos del Reporte:</h2>
                    <pre style="white-space: pre-wrap; font-size: 12px;">{json.dumps(report.data, indent=2, ensure_ascii=False)}</pre>
                </div>
                
                <div class="footer">
                    <p>SmartSales - Sistema de Reportes Inteligente</p>
                    <p>Generado automáticamente el {timezone.now().strftime('%d/%m/%Y %H:%M')}</p>
                </div>
            </body>
            </html>
            """
            
            if WEASYPRINT_AVAILABLE:
                # Usar WeasyPrint para generar PDF
                pdf_file = HTML(string=html_content).write_pdf()
                
                response = HttpResponse(pdf_file, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="reporte_{report.id}_{timezone.now().strftime("%Y%m%d")}.pdf"'
                return response
            else:
                # Fallback: devolver HTML
                response = HttpResponse(html_content, content_type='text/html')
                response['Content-Disposition'] = f'attachment; filename="reporte_{report.id}_{timezone.now().strftime("%Y%m%d")}.html"'
                return response
                
        except Exception as e:
            return Response({'error': f'Error al generar PDF: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def export_excel(self, request):
        """Exportar reporte a Excel"""
        try:
            report_id = request.data.get('report_id')
            if not report_id:
                return Response({'error': 'report_id es requerido'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                report = GeneratedReport.objects.get(id=report_id)
            except GeneratedReport.DoesNotExist:
                return Response({'error': 'Reporte no encontrado'}, status=status.HTTP_404_NOT_FOUND)
            
            if EXCEL_AVAILABLE:
                # Crear archivo Excel con openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Reporte"
                
                # Encabezado
                ws['A1'] = report.title
                ws['A2'] = f"Generado: {report.created_at.strftime('%d/%m/%Y %H:%M')}"
                ws['A3'] = f"Estado: {report.status}"
                
                # Datos
                row = 5
                if isinstance(report.data, dict):
                    for key, value in report.data.items():
                        ws[f'A{row}'] = str(key)
                        ws[f'B{row}'] = str(value)
                        row += 1
                elif isinstance(report.data, list):
                    if report.data and isinstance(report.data[0], dict):
                        # Escribir encabezados
                        headers = list(report.data[0].keys())
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=header)
                        row += 1
                        
                        # Escribir datos
                        for item in report.data:
                            for col, header in enumerate(headers, 1):
                                ws.cell(row=row, column=col, value=str(item.get(header, '')))
                            row += 1
                
                # Guardar en memoria
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(
                    output.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="reporte_{report.id}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
                return response
            else:
                # Fallback: generar CSV
                output = io.StringIO()
                writer = csv.writer(output)
                
                # Escribir encabezado
                writer.writerow([report.title])
                writer.writerow([f"Generado: {report.created_at.strftime('%d/%m/%Y %H:%M')}"])
                writer.writerow([f"Estado: {report.status}"])
                writer.writerow([])  # Línea vacía
                
                # Escribir datos
                if isinstance(report.data, dict):
                    writer.writerow(['Campo', 'Valor'])
                    for key, value in report.data.items():
                        writer.writerow([key, value])
                elif isinstance(report.data, list):
                    if report.data and isinstance(report.data[0], dict):
                        headers = list(report.data[0].keys())
                        writer.writerow(headers)
                        for item in report.data:
                            writer.writerow([item.get(header, '') for header in headers])
                
                response = HttpResponse(output.getvalue(), content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="reporte_{report.id}_{timezone.now().strftime("%Y%m%d")}.csv"'
                return response
                
        except Exception as e:
            return Response({'error': f'Error al generar Excel: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
